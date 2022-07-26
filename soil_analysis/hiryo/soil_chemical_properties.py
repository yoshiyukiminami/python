# Step-1 soil_chemical_propertiesフォルダにある測定データ（xlsx）を読み込む
# Step-2 土壌化学性分析データを抽出し、グラフ表示用のDATAFRAMEに取り込む
# Step-3 グラフ化する
import pandas as pd
import datetime
import glob
import os
import pprint
import openpyxl

def dataappend_dik5531(df, df_file1, df_file2, diktype, datanumber):
    # print(df, df_file1, df_file2)
    depth = df_file1.iat[3, 1]
    spring = df_file1.iat[5, 1]
    cone = df_file1.iat[6, 1]
    # print(depth, spring, cone)
    koudodata = df_file2.iloc[:, 1]
    # print(koudodata, type(koudodata))

    isvalid = True
    if not depth == 60 and spring == 490 and cone == 2:
        print("硬度計の設定が間違っている可能性があります")
        isvalid = False
    else:
        print("ここまで途中")
        # for data in df_file2:
        #     df.concat()


def dataappend_dik5532(df, df_file3, df_file4, diktype, datanumber2):
    # print(df, df_file3, df_file4)
    depth = df_file3.iat[3, 1]
    cone = df_file3.iat[5, 1]
    angle = df_file3.iat[9, 1]
    print(depth, angle, cone)
    koudodata = df_file4.iloc[:, 1]
    # print(koudodata, type(koudodata))

    isvalid = True
    if not depth == 60 and angle == 30 and cone == 2:
        print("硬度計の設定が間違っている可能性があります")
        isvalid = False
    else:
        print("ここまで途中")
        # for data in df_file2:
        #     df.concat()


if __name__ == '__main__':
    # Step-1 フォルダにある測定データ（xlsx）を読み込む
    # 読み込むデータを特定するfolder-pathをfiledirに設定する
    filedir = 'C:/Users/minam/Desktop/soil_chemical_properties/'
    # フォルダー内にあるフォルダー名をfolderlist、ファイル名をfilesに所得する
    folderlist = os.listdir(filedir)
    print(folderlist)
    files = glob.glob(filedir + '/**/*.xlsx', recursive=True)
    pprint.pprint(files)

    # フォルダにある測定データ（.xlsx）を読み込む
    for file in files:
        df = pd.ExcelFile(file)
        print(df.sheet_names)
        df2 = openpyxl.load_workbook(file)
        df_chem = df2['土壌化学性データ']
        print(type(df_chem))
        cell1 = df_chem.cell(2, 1)
        print(cell1)
        cell2 = df_chem.cell(2, 2)
        print(cell2)

    # for file in files:
    #     # Step-2 地点土壌硬度データから硬度計タイプを判定する
    #     Nstr = len(file)
    #     diktype = file[Nstr-40:Nstr-32]
    #     print(diktype)
    #     if diktype == 'DIK-5531':
    #         print("硬度計のタイプは旧型（5531）です")
    #         datanumber = file[Nstr-31:Nstr-27]
    #         print(datanumber)
    #         df_file1 = pd.read_csv(file, encoding='Shift-JIS', header=0, nrows=7)
    #         df_file2 = pd.read_csv(file, encoding='Shift-JIS', header=0, skiprows=9)
    #         dataappend_dik5531(df, df_file1, df_file2, diktype, datanumber)
    #     else:
    #         print("硬度計のタイプは新型（5532）です")
    #         dataforder = file[Nstr-9:Nstr-8]
    #         print(dataforder)
    #         datanumber2 = dataforder + '-' + file[Nstr-7:Nstr-3]
    #         print(datanumber2)
    #         df_file3 = pd.read_csv(file, encoding='Shift-JIS', header=0, nrows=10)
    #         df_file4 = pd.read_csv(file, encoding='Shift-JIS', header=0, skiprows=11)
    #         diktype = 'DIK-5532'
    #         dataappend_dik5532(df, df_file3, df_file4, diktype, datanumber2)

